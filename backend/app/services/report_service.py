"""
IntelliDesk AI — Report Generation Service
Generates PDF reports (ReportLab) and Excel/CSV exports (openpyxl/pandas).
"""

import io
import csv
from datetime import datetime, timezone, timedelta, date
from typing import Optional

from app.utils.logger import get_logger

logger = get_logger(__name__)


class ReportService:
    """
    Generates downloadable reports in PDF, CSV, and Excel formats.
    All methods return BytesIO objects for direct HTTP response streaming.
    """

    # ─── PDF Report ───────────────────────────────────────────────────────────

    @staticmethod
    def generate_ticket_report_pdf(
        from_date: date,
        to_date: date,
        department_id: Optional[int] = None,
    ) -> io.BytesIO:
        """
        Generate a styled PDF report for tickets in a date range.
        Includes: summary table, by-status breakdown, SLA compliance, top-5 agents.
        """
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable,
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        from app.models.ticket import Ticket
        from app.models.user import User
        from app.extensions import db
        from app.utils.constants import TicketStatus
        from sqlalchemy import func

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        brand_blue = colors.HexColor("#1E40AF")
        brand_light = colors.HexColor("#EFF6FF")

        title_style = ParagraphStyle(
            "title", parent=styles["Heading1"],
            textColor=brand_blue, fontSize=20, alignment=TA_CENTER, spaceAfter=6,
        )
        subtitle_style = ParagraphStyle(
            "subtitle", parent=styles["Normal"],
            textColor=colors.HexColor("#6B7280"), fontSize=10, alignment=TA_CENTER, spaceAfter=20,
        )
        section_style = ParagraphStyle(
            "section", parent=styles["Heading2"],
            textColor=brand_blue, fontSize=13, spaceBefore=16, spaceAfter=8,
        )

        elements = []

        # Header
        elements.append(Paragraph("IntelliDesk AI", title_style))
        elements.append(Paragraph("Ticket Performance Report", styles["Heading2"]))
        elements.append(Paragraph(
            f"Period: {from_date.strftime('%d %b %Y')} – {to_date.strftime('%d %b %Y')}  |  "
            f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}",
            subtitle_style,
        ))
        elements.append(HRFlowable(width="100%", thickness=1, color=brand_blue))
        elements.append(Spacer(1, 12))

        # ── Ticket summary ────────────────────────────────────────────────────
        start = datetime.combine(from_date, datetime.min.time()).replace(tzinfo=timezone.utc)
        end = datetime.combine(to_date, datetime.max.time()).replace(tzinfo=timezone.utc)

        base_q = Ticket.query.filter(
            Ticket.created_at.between(start, end),
            Ticket.deleted_at.is_(None),
        )
        if department_id:
            base_q = base_q.filter(Ticket.department_id == department_id)

        total = base_q.count()
        resolved = base_q.filter(Ticket.status == TicketStatus.RESOLVED.value).count()
        closed = base_q.filter(Ticket.status == TicketStatus.CLOSED.value).count()
        open_count = base_q.filter(
            Ticket.status.notin_([TicketStatus.RESOLVED.value, TicketStatus.CLOSED.value])
        ).count()
        sla_breached = base_q.filter(Ticket.sla_resolution_breached == True).count()  # noqa: E712
        sla_rate = round((total - sla_breached) / total * 100, 1) if total > 0 else 100.0

        summary_data = [
            ["Metric", "Value"],
            ["Total Tickets Created", str(total)],
            ["Resolved", str(resolved)],
            ["Closed", str(closed)],
            ["Open / In-Progress", str(open_count)],
            ["SLA Breached", str(sla_breached)],
            ["SLA Compliance Rate", f"{sla_rate}%"],
        ]

        elements.append(Paragraph("1. Ticket Summary", section_style))
        summary_table = Table(summary_data, colWidths=[10 * cm, 6 * cm])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), brand_blue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [brand_light, colors.white]),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
            ("PADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 16))

        # ── By-status breakdown ───────────────────────────────────────────────
        elements.append(Paragraph("2. Breakdown by Status", section_style))
        status_rows = (
            db.session.query(Ticket.status, func.count(Ticket.id).label("cnt"))
            .filter(Ticket.created_at.between(start, end), Ticket.deleted_at.is_(None))
            .group_by(Ticket.status)
            .all()
        )
        status_data = [["Status", "Count"]] + [
            [row.status.replace("_", " ").title(), str(row.cnt)] for row in status_rows
        ]
        status_table = Table(status_data, colWidths=[10 * cm, 6 * cm])
        status_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [brand_light, colors.white]),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
            ("PADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(status_table)
        elements.append(Spacer(1, 16))

        # ── Priority breakdown ────────────────────────────────────────────────
        elements.append(Paragraph("3. Breakdown by Priority", section_style))
        priority_rows = (
            db.session.query(Ticket.priority, func.count(Ticket.id).label("cnt"))
            .filter(Ticket.created_at.between(start, end), Ticket.deleted_at.is_(None))
            .group_by(Ticket.priority)
            .all()
        )
        priority_data = [["Priority", "Count"]] + [
            [(r.priority or "Unknown").title(), str(r.cnt)] for r in priority_rows
        ]
        priority_table = Table(priority_data, colWidths=[10 * cm, 6 * cm])
        priority_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [brand_light, colors.white]),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
            ("PADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(priority_table)

        # Footer
        elements.append(Spacer(1, 30))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CBD5E1")))
        elements.append(Paragraph(
            "Confidential — Generated by IntelliDesk AI Platform",
            ParagraphStyle("footer", parent=styles["Normal"],
                           fontSize=8, textColor=colors.HexColor("#9CA3AF"), alignment=TA_CENTER),
        ))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    # ─── CSV Export ───────────────────────────────────────────────────────────

    @staticmethod
    def export_tickets_csv(
        from_date: date,
        to_date: date,
        department_id: Optional[int] = None,
    ) -> io.BytesIO:
        """
        Export tickets in a date range as a UTF-8 CSV file.
        Includes: ticket number, title, status, priority, requester, assignee, SLA, dates.
        """
        from app.models.ticket import Ticket

        start = datetime.combine(from_date, datetime.min.time()).replace(tzinfo=timezone.utc)
        end = datetime.combine(to_date, datetime.max.time()).replace(tzinfo=timezone.utc)

        query = Ticket.query.filter(
            Ticket.created_at.between(start, end),
            Ticket.deleted_at.is_(None),
        )
        if department_id:
            query = query.filter(Ticket.department_id == department_id)

        tickets = query.order_by(Ticket.created_at.asc()).all()

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([
            "Ticket Number", "Title", "Status", "Priority", "Category",
            "Requester", "Assignee", "Department",
            "SLA Response Deadline", "SLA Resolution Deadline",
            "SLA Breached", "First Responded At", "Resolved At",
            "Created At", "Updated At",
        ])

        for t in tickets:
            req = f"{t.requester.first_name} {t.requester.last_name}" if t.requester else ""
            asgn = f"{t.assignee.first_name} {t.assignee.last_name}" if t.assignee else ""
            dept = t.department.name if t.department else ""
            writer.writerow([
                t.ticket_number, t.title, t.status,
                t.priority or "", t.category or "",
                req, asgn, dept,
                t.sla_response_deadline.isoformat() if t.sla_response_deadline else "",
                t.sla_resolution_deadline.isoformat() if t.sla_resolution_deadline else "",
                "Yes" if t.sla_resolution_breached else "No",
                t.first_responded_at.isoformat() if t.first_responded_at else "",
                t.resolved_at.isoformat() if t.resolved_at else "",
                t.created_at.isoformat(),
                t.updated_at.isoformat(),
            ])

        csv_bytes = buffer.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
        return io.BytesIO(csv_bytes)

    # ─── Excel Export ─────────────────────────────────────────────────────────

    @staticmethod
    def export_analytics_excel(
        from_date: date,
        to_date: date,
    ) -> io.BytesIO:
        """
        Generate a multi-sheet Excel analytics workbook.
        Sheets: Daily Trends, SLA Compliance, Agent Performance.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from app.services.analytics_service import AnalyticsService
        from app.models.analytics import DailyMetricSnapshot, AgentDailyMetric

        wb = openpyxl.Workbook()
        brand_fill = PatternFill("solid", fgColor="1E40AF")
        white_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        def style_header_row(ws, headers: list[str], row: int = 1):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = brand_fill
                cell.font = white_font
                cell.alignment = header_align
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col)].width = max(14, len(header) + 4)

        # ── Sheet 1: Daily Trends ─────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Daily Trends"
        days = (to_date - from_date).days + 1
        trends = AnalyticsService.get_trend_data(days=days)

        headers1 = [
            "Date", "Tickets Created", "Tickets Resolved", "Open",
            "Overdue", "SLA Compliance %", "Avg Resolution (hrs)",
            "Incidents Created", "Critical Incidents",
        ]
        style_header_row(ws1, headers1)
        for row_idx, row_data in enumerate(trends, 2):
            ws1.append([
                row_data["date"],
                row_data["tickets_created"],
                row_data["tickets_resolved"],
                row_data["tickets_open"],
                row_data["tickets_overdue"],
                row_data["sla_compliance_rate"],
                row_data["avg_resolution_hours"],
                row_data["incidents_created"],
                row_data["critical_incidents"],
            ])

        # ── Sheet 2: SLA Compliance ────────────────────────────────────────────
        ws2 = wb.create_sheet("SLA Compliance")
        headers2 = ["Priority", "Total Tickets", "SLA Met", "SLA Breached", "Compliance Rate %"]
        style_header_row(ws2, headers2)
        sla = AnalyticsService.get_sla_compliance_by_priority()
        for priority, data in sla.items():
            ws2.append([
                priority.title(),
                data["total"],
                data["compliant"],
                data["breached"],
                data["compliance_rate"],
            ])

        # ── Sheet 3: Agent Performance ─────────────────────────────────────────
        ws3 = wb.create_sheet("Agent Performance")
        headers3 = [
            "Agent Name", "Tickets Resolved", "Tickets Assigned",
            "Resolution Rate %", "SLA Breached", "Avg Resolution (hrs)"
        ]
        style_header_row(ws3, headers3)
        leaderboard = AnalyticsService.get_agent_leaderboard(days=(to_date - from_date).days + 1)
        for agent in leaderboard:
            ws3.append([
                agent["agent_name"],
                agent["tickets_resolved"],
                agent["tickets_assigned"],
                agent["resolution_rate"],
                agent["sla_breached"],
                agent["avg_resolution_hours"],
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
